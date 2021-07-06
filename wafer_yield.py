import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class Die:
    def __init__(self, wafer: str, bin: int, sub_bin: int, reading_1: float, reading_2: float):
        self.wafer = wafer
        self.bin = bin
        self.sub_bin = sub_bin
        self.reading_1 = reading_1
        self.reading_2 = reading_2

    def __repr__(self):
        return f"Die({self.wafer}, {self.bin}, {self.sub_bin}, {self.reading_1}, {self.reading_2})"


class RawData:
    _cols = ["Wafer", "Bin", "Sub_Bin", "Reading_1", "Reading_2"]
    def __init__(self, sheet=None):
        self.dies = {}
        self.bin_hierarchy = {}
        if sheet != None:
            self.load_raw_data_sheet(sheet)

    def __repr__(self):
        return f"RawData({self.dies})"

    def create_all_df(self, wb):
        for func in (self.create_df1, self.create_df1_1, self.create_df2, self.create_df2_1, self.create_df3, self.create_df3_1):
            func(wb)

    def load_raw_data_sheet(self, sheet: Worksheet) -> None:
        def get_values(cells):
            return [cell.value for cell in cells]
        rows = iter(sheet.rows)
        assert (header := get_values(next(rows))) == self._cols
        for row in rows:
            die = Die(*get_values(row))
            try:
                wafer = self.dies[die.wafer]
                try:
                    bin = wafer[die.bin]
                    try:
                        bin[die.sub_bin].append(die)
                    except KeyError:
                        bin[die.sub_bin] = [die]
                except KeyError:
                    wafer[die.bin] = {die.sub_bin: [die]}
            except KeyError:
                self.dies[die.wafer] = {die.bin: {die.sub_bin: [die]}}
            try:
                self.bin_hierarchy[die.bin].add(die.sub_bin)
            except KeyError:
                self.bin_hierarchy[die.bin] = set()
                self.bin_hierarchy[die.bin].add(die.sub_bin)

    @staticmethod
    def generate_sheet_name(basename) -> str:
        return basename + "*"

    @staticmethod
    def is_generated_sheet_name(sheetname) -> bool:
        return sheetname.endswith('*')

    @staticmethod
    def to_percentage(val: float) -> str:
        return f"{val:.2f}%"
        
    @staticmethod
    def get_sheet(wb, sheetname) -> Worksheet:
        if sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            sheet.delete_rows(0, sheet.max_row)
        else:
            sheet = wb.create_sheet(sheetname, len(wb.worksheets))
        return sheet

    def get_bin_lut(self):
        bin_names = sorted(self.bin_hierarchy)
        return bin_names, {v:i for i, v in enumerate(bin_names)}

    def get_sub_bin_lut(self):
        sub_bin_names = [sub_bin_name for bin_name in sorted(self.bin_hierarchy) for sub_bin_name in sorted(self.bin_hierarchy[bin_name])]
        lut = {v:i for i, v in enumerate(sub_bin_names)}
        idx, bin_names, bin_offsets = 0, [], []
        for bin_name in sorted(self.bin_hierarchy):
            bin_names.append(bin_name)
            bin_offsets.append(idx)
            idx += len(self.bin_hierarchy[bin_name])
        return sub_bin_names, lut, bin_names, bin_offsets

    def create_df1(self, wb, sheetname="df1-gen", percentage=False):
        sheet = self.get_sheet(wb, sheetname)
        bin_names, lut = self.get_bin_lut()
        for row in ([], [], ["計數 - Bin", "Bin"], ["Wafer"] + bin_names + ["總計"]):
            sheet.append(row)
        for wafer_name in sorted(self.dies):
            results = [0 for each in bin_names]
            for bin_name, bins in self.dies[wafer_name].items():
                results[lut[bin_name]] = sum(map(len, bins.values()))
            results.append(cnt:=sum(results))
            if percentage:
                results = [self.to_percentage(each / cnt) for each in results]
            sheet.append([wafer_name] + results)

    def create_df1_1(self, wb, sheetname="df1-1-gen"):
        return self.create_df1(wb, sheetname, True)

    def create_df2(self, wb, sheetname="df2-gen", percentage=False):
        sheet = self.get_sheet(wb, sheetname)
        sub_bin_names, lut, bin_names, bin_offsets = self.get_sub_bin_lut()
        for row in ([], [], ["計數 - Bin", "Bin", "Sub_Bin"], [], ["Wafer"] + sub_bin_names + ["總計"]):
            sheet.append(row)
        for bin_name, offset in zip(bin_names, bin_offsets):
            sheet.cell(row=4, column=offset+1).value = bin_name
        for wafer_name in sorted(self.dies):
            results = [0 for each in sub_bin_names]
            for bin_dies in (dies:=self.dies[wafer_name].values()):
                for sub_bin_name, sub_bin_dies in bin_dies.items():
                    results[lut[sub_bin_name]] = len(sub_bin_dies)
            results.append(cnt:=sum(results))
            if percentage:
                results = [self.to_percentage(each / cnt) for each in results]
            sheet.append([wafer_name] + results)

    def create_df2_1(self, wb, sheetname="df2-1-gen"):
        return self.create_df2(wb, sheetname, True)

    def create_df3(self, wb, sheetname="df3-gen"):
        sheet = self.get_sheet(wb, sheetname)
        for row in ([], [], ["Wafer", "Bin", "Sub_Bin", "平均值 - Reading_1", "最大 - Reading_2"]):
            sheet.append(row)
        for wafer_name in sorted(self.dies):
            for bin_name in sorted(self.bin_hierarchy):
                try:
                    dies = [die for sub_bin_dies in self.dies[wafer_name][bin_name].values() for die in sub_bin_dies]
                    sheet.append([wafer_name, bin_name, sum((die.reading_1 for die in dies))/len(dies), max(die.reading_2 for die in dies)])
                except (KeyError, ZeroDivisionError):
                    continue
                    # sheet.append([wafer_name, bin_name, None, None])
                    
    def create_df3_1(self, wb, sheetname="df3-1-gen"):
        sheet = self.get_sheet(wb, sheetname)
        for row in ([], [], ["Wafer", "Bin", "Sub_Bin", "平均值 - Reading_1", "最大 - Reading_2"]):
            sheet.append(row)
        for wafer_name in sorted(self.dies):
            for bin_name in sorted(self.dies[wafer_name]):
                for sub_bin_name in sorted(self.dies[wafer_name][bin_name]):
                    sub_bin_dies = self.dies[wafer_name][bin_name][sub_bin_name]
                    try:
                        sheet.append([wafer_name, bin_name, sub_bin_name, sum((die.reading_1 for die in sub_bin_dies))/len(sub_bin_dies), max(die.reading_2 for die in sub_bin_dies)])
                    except (KeyError, ZeroDivisionError):
                        continue
                        # sheet.append([wafer_name, bin_name, None, None])


def test_sample(filename=r"Wafer Yield.xlsx", sheetname="Raw Data", outfilename="out.xlsx"):
    wb = openpyxl.load_workbook(filename)
    raw_data = RawData(wb[sheetname])
    raw_data.create_all_df(wb)
    wb.save(outfilename)


if __name__ == "__main__":
    test_sample()